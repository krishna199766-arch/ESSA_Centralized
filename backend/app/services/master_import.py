"""Reading a master out of a file somebody already keeps.

WHY THIS EXISTS. A catalogue's categories and its attribute vocabularies are
hundreds of rows, and they already exist — in a spreadsheet on somebody's
machine, or printed. Typing them into a web form is an afternoon's work that
produces typos, and it is the single thing standing between a new business line
and being usable. So the file is read.

WHAT IT READS.

  .xlsx / .xlsm  the real answer. A spreadsheet has cells, so the columns are
                 unambiguous and nothing has to be guessed.
  .csv / .tsv    the same, with no dependency at all.
  .pdf           text only, one value per line. Honest about its limits: a PDF
                 has no cells, only glyphs at coordinates, and a laid-out table
                 comes back as run-together lines. It is offered because a list
                 printed as a PDF is a real thing people have, and refusing it
                 outright would send them to retype it — but a PDF import is
                 always shown for review before anything is written.

WHY EVERY IMPORT PREVIEWS FIRST. The router runs these with `commit=False` by
default and hands back exactly what WOULD be written. A file that silently added
four hundred wrong values to a live master would be discovered weeks later, by
which time products would be classified against them. Reading a file is a guess
about someone's layout; showing the guess is what makes it safe.

TWO SHAPES ARE ACCEPTED, because both are what people actually have:

  WIDE — one column per attribute, its values beneath it. This is the shape of
         Essa's own Attributes Reference.xlsx.

             Colour | Size | Weave
             White  | S    | Kanchipuram
             Black  | M    | Banarasi

  LONG — one row per value, naming its attribute.

             Attribute | Value
             Colour    | White
             Colour    | Black
             Weave     | Kanchipuram
"""
import csv
import io
import re

#: Cells this long are prose, not a master value — a paragraph that wandered into
#: a spreadsheet, or a PDF line that is really a sentence. Dropped rather than
#: stored, because a dropdown entry nobody can read is worse than a missing one.
MAX_VALUE = 80

#: Headers that mean "this column names the attribute" / "this column is a value",
#: which is what tells the LONG shape from the WIDE one.
_ATTR_HEADERS = {"attribute", "attribute name", "attr", "field", "property",
                 "key", "name"}
_VALUE_HEADERS = {"value", "values", "option", "options", "entry"}
_SECTION_HEADERS = {"section", "group", "gender", "division"}
_CATEGORY_HEADERS = {"category", "categories", "category name", "name", "product"}

#: Friendly spellings people actually type, mapped to the keys the app uses. A
#: sheet headed "Colour" must land on `color` — the column the whole app reads —
#: rather than creating a second attribute that means the same thing.
HEADER_ALIASES = {
    "colour": "color", "color": "color",
    "size": "size", "sizes": "size",
    "type": "product_type", "product type": "product_type",
    "producttype": "product_type",
    "material": "material", "fabric": "material",
    "pattern": "pattern", "fit": "fit", "style": "style",
    "sleeve": "sleeve", "sleeves": "sleeve",
    "brand": "brand",
    "design": "design_no", "design no": "design_no", "designno": "design_no",
    "design number": "design_no",
}


class ImportError_(ValueError):
    """A file this module cannot make sense of, said in words a user can act on."""


# ---------------------------------------------------------------------------
#  file → rows
# ---------------------------------------------------------------------------
def _clean(cell):
    if cell is None:
        return ""
    text = str(cell).strip()
    # Excel hands back 12.0 for a cell someone typed 12 into; a size of "12.0"
    # is not a size anybody will match against.
    if re.fullmatch(r"-?\d+\.0", text):
        text = text[:-2]
    return re.sub(r"\s+", " ", text)


def _rows_from_xlsx(data: bytes) -> list:
    try:
        import openpyxl
    except ImportError:                       # pragma: no cover - deployment guard
        raise ImportError_("this server cannot read .xlsx files — save the sheet "
                           "as CSV and upload that instead")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:                  # noqa: BLE001 — reported, not raised
        raise ImportError_(f"that file could not be opened as a spreadsheet ({exc})")
    try:
        # The first sheet only. A workbook's later sheets are usually notes,
        # pivot tables or last year's copy, and importing all of them silently is
        # how a master doubles in size.
        ws = wb[wb.sheetnames[0]]
        return [[_clean(c) for c in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _rows_from_csv(data: bytes) -> list:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [[_clean(c) for c in row] for row in csv.reader(io.StringIO(text), dialect)]


def _rows_from_pdf(data: bytes) -> list:
    try:
        from pypdf import PdfReader
    except ImportError:                       # pragma: no cover - deployment guard
        raise ImportError_("this server cannot read PDFs — save the list as CSV "
                           "or Excel instead")
    try:
        reader = PdfReader(io.BytesIO(data))
        text = "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:                  # noqa: BLE001
        raise ImportError_(f"that PDF could not be read ({exc})")
    if not text.strip():
        raise ImportError_("that PDF has no text in it — it is probably a scan. "
                           "Upload the list as CSV or Excel instead.")
    # One value per line. A PDF has no cells, so anything more structured than
    # this would be guessing at column boundaries from whitespace — which is
    # exactly the guess that produces a master full of half-values.
    return [[_clean(line)] for line in text.splitlines() if _clean(line)]


def read_rows(filename: str, data: bytes) -> list:
    """The file as a grid of strings. Raises ImportError_ with a readable reason."""
    name = (filename or "").lower()
    if not data:
        raise ImportError_("that file is empty")
    if name.endswith((".xlsx", ".xlsm")):
        rows = _rows_from_xlsx(data)
    elif name.endswith((".csv", ".tsv", ".txt")):
        rows = _rows_from_csv(data)
    elif name.endswith(".pdf"):
        rows = _rows_from_pdf(data)
    elif name.endswith(".xls"):
        raise ImportError_("the old .xls format can't be read — open it in Excel "
                           "and save as .xlsx or CSV")
    else:
        raise ImportError_("upload a .xlsx, .csv or .pdf file")
    rows = [r for r in rows if any(c for c in r)]
    if not rows:
        raise ImportError_("there are no rows in that file")
    return rows


# ---------------------------------------------------------------------------
#  rows → meaning
# ---------------------------------------------------------------------------
def normalise_key(name: str) -> str:
    """A header as the app's attribute key. "Design No" → design_no."""
    text = _clean(name).lower()
    if text in HEADER_ALIASES:
        return HEADER_ALIASES[text]
    key = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return key


def _looks_like_header(row, rest) -> bool:
    """Whether the first row names the columns rather than being data.

    Decided by comparing it to what follows: a header is short text in every
    filled cell, and the rows under it are not identical in shape to it. With
    only one row in the file there is nothing to compare against, so it is
    treated as data — losing the single value somebody uploaded would be worse
    than importing a stray header they can delete.
    """
    filled = [c for c in row if c]
    if not filled or not rest:
        return False
    if any(len(c) > 40 for c in filled):
        return False
    known = sum(1 for c in filled
                if normalise_key(c) in HEADER_ALIASES.values()
                or _clean(c).lower() in (_ATTR_HEADERS | _VALUE_HEADERS
                                         | _SECTION_HEADERS | _CATEGORY_HEADERS))
    # A row naming things the app already knows is a header, near enough.
    return known >= max(1, len(filled) // 2)


def _column(rows, index):
    return [r[index] for r in rows if index < len(r) and r[index]]


def parse_attributes(rows) -> dict:
    """{key: {"label": …, "values": [...]}} from a WIDE or LONG sheet.

    The shape is detected, not configured, because asking somebody which layout
    their file is in requires them to know what the layouts are called.
    """
    header = rows[0]
    body = rows[1:]
    has_header = _looks_like_header(header, body)
    if not has_header:
        raise ImportError_(
            "the first row has to name the attributes — put a heading over each "
            "column (Colour, Size, Weave…), or use two columns headed "
            "Attribute and Value")

    labels = [c for c in header if c]
    lowered = [_clean(c).lower() for c in header]

    # LONG: exactly two named columns, one saying which attribute and one the value
    if len(labels) == 2 and lowered[0] in _ATTR_HEADERS and lowered[1] in _VALUE_HEADERS:
        found = {}
        for row in body:
            if len(row) < 2:
                continue
            attr, value = _clean(row[0]), _clean(row[1])
            if not attr or not value or len(value) > MAX_VALUE:
                continue
            key = normalise_key(attr)
            if not key:
                continue
            entry = found.setdefault(key, {"label": attr, "values": []})
            entry["values"].append(value)
        return _dedupe(found)

    # WIDE: one column per attribute
    found = {}
    for i, name in enumerate(header):
        if not name:
            continue
        key = normalise_key(name)
        if not key:
            continue
        values = [v for v in _column(body, i) if len(v) <= MAX_VALUE]
        found.setdefault(key, {"label": _clean(name), "values": []})["values"].extend(values)
    if not found:
        raise ImportError_("no attribute columns were found in that file")
    return _dedupe(found)


def _dedupe(found: dict) -> dict:
    """Fold repeats case-insensitively, keeping the first spelling and the order.

    The order is the file's, not alphabetical: a master usually arrives sorted by
    how often each value occurs, and re-sorting it buries the common ones.
    """
    out = {}
    for key, entry in found.items():
        seen, values = set(), []
        for v in entry["values"]:
            k = v.lower()
            if k in seen:
                continue
            seen.add(k)
            values.append(v)
        out[key] = {"label": entry["label"], "values": values}
    return out


def parse_values(rows, attr_key=None, attr_label=None) -> list:
    """A flat list of values for ONE attribute — the first column of the file.

    `attr_key` / `attr_label` are what this list is being imported INTO, and they
    are what makes the header check work here. A one-column export is headed with
    the attribute's own name — a file of zari values starts with the word "Zari"
    — and that word is not a generic header like "Value", so nothing but knowing
    the destination can tell it from a value. Without this, every such import
    quietly added the attribute's own name to its dropdown.
    """
    first = _clean(rows[0][0]) if rows and rows[0] else ""
    names = {n for n in (attr_key, attr_label) if n}
    names |= {normalise_key(n) for n in names}
    is_own_name = bool(first) and (first.lower() in {str(n).lower() for n in names}
                                   or normalise_key(first) in names)
    body = (rows[1:] if is_own_name or _looks_like_header(rows[0], rows[1:])
            else rows)
    seen, out = set(), []
    for row in body:
        for cell in row:
            v = _clean(cell)
            if not v or len(v) > MAX_VALUE:
                continue
            if v.lower() in seen:
                continue
            seen.add(v.lower())
            out.append(v)
            break                      # first filled cell in the row, and no more
    if not out:
        raise ImportError_("no values were found in that file")
    return out


# ---------------------------------------------------------------------------
#  meaning → file
# ---------------------------------------------------------------------------
#  The export lives beside the import ON PURPOSE. What comes out has to be a file
#  this module can read back in, because the useful thing to do with a master is
#  download it, edit it in Excel, and upload it again. Keeping the two halves in
#  one file is what stops the headers drifting apart — a renamed heading here is
#  a heading the parser above stops recognising, and the round trip breaks
#  silently.
# ---------------------------------------------------------------------------
def to_csv(rows) -> bytes:
    """A grid as CSV, BOM-prefixed so Excel opens it as UTF-8.

    Without the BOM, Excel on Windows reads a CSV as the system codepage and a
    rupee sign or an accented colour name comes out as mojibake — on the one
    machine this file is most likely to be opened on.
    """
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    for row in rows:
        writer.writerow(["" if c is None else str(c) for c in row])
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def to_xlsx(rows, sheet_title="Master") -> bytes:
    try:
        import openpyxl
    except ImportError:                       # pragma: no cover - deployment guard
        raise ImportError_("this server cannot write .xlsx files — download CSV instead")
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Master"
    for row in rows:
        ws.append(["" if c is None else str(c) for c in row])
    if rows:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # Frozen so the headings stay visible while somebody scrolls four hundred
        # colours — this file exists to be edited, not just looked at.
        ws.freeze_panes = "A2"
        for i, _ in enumerate(rows[0], start=1):
            letter = ws.cell(row=1, column=i).column_letter
            longest = max((len(str(r[i - 1])) for r in rows
                           if i - 1 < len(r) and r[i - 1]), default=8)
            ws.column_dimensions[letter].width = min(40, max(12, longest + 2))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def attribute_grid(attributes, options) -> list:
    """The WIDE grid: a column per attribute, its values beneath.

    Exactly the shape `parse_attributes` reads, so a download can be edited and
    uploaded straight back. Attributes with no values still get their column —
    that column IS the template for filling them in.
    """
    if not attributes:
        return [["Attribute", "Value"]]
    header = [a["label"] for a in attributes]
    columns = [list(options.get(a["key"], [])) for a in attributes]
    depth = max((len(c) for c in columns), default=0)
    grid = [header]
    for i in range(depth):
        grid.append([c[i] if i < len(c) else "" for c in columns])
    return grid


def values_grid(label, values) -> list:
    """One attribute's list, headed with its own name — which `parse_values`
    recognises as a heading rather than importing as a value."""
    return [[label]] + [[v] for v in values]


def category_grid(categories) -> list:
    return [["Category", "Section"]] + [[c.name, c.section or "OVERALL"]
                                        for c in categories]


def parse_categories(rows) -> list:
    """[{"name":…, "section":…}] from a file of category names.

    A `Section` column is honoured where there is one — the garment master splits
    OVERALL / KIDS / LADIES / MENS and the section is what a report groups by.
    Without one everything lands in OVERALL, which is what the master's own
    unsectioned rows already are.
    """
    header = rows[0]
    has_header = _looks_like_header(header, rows[1:])
    body = rows[1:] if has_header else rows

    name_col, section_col = 0, None
    if has_header:
        for i, cell in enumerate(header):
            low = _clean(cell).lower()
            if low in _CATEGORY_HEADERS and name_col == 0:
                name_col = i
            elif low in _SECTION_HEADERS:
                section_col = i

    seen, out = set(), []
    for row in body:
        name = _clean(row[name_col]) if name_col < len(row) else ""
        if not name or len(name) > MAX_VALUE:
            continue
        name = name.upper()
        if name.lower() in seen:
            continue
        seen.add(name.lower())
        section = ""
        if section_col is not None and section_col < len(row):
            section = _clean(row[section_col]).upper()
        out.append({"name": name, "section": section or "OVERALL"})
    if not out:
        raise ImportError_("no category names were found in that file")
    return out
